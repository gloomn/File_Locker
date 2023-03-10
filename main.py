# © 2022 Lee-Ki-Joon(Gloomn) <ithan0704@naver.com>

from tkinter import *
import tkinter as tk
import tkinter.messagebox as tkmb
from glui import standardButton
from glui import standardLabel
from glui import standardEntry
import subprocess
import os
import hashlib
import base64
import chardet
from registerForm import *
import bcrypt
from openpyxl import load_workbook
import keyboard
import subprocess
#window default setting
window = tk.Tk()
window.configure(bg="#fff")
window.title("File Locker (Test Version: alphaV)")
window.geometry("600x300")
window.resizable(False, False)
window.eval('tk::PlaceWindow . center')

#login variables
luser_id, lpassword = StringVar(window), StringVar(window)

def register():
    load_wb = load_workbook("#20$6&7!#04.xlsx", data_only=True)
    load_ws = load_wb['asdfas9df72']
    if(load_ws['A1'].value == 1 and load_ws['A2'].value == 1):
        registerFormAction()
    else:
        tkmb.showerror("Error Code: 005", "Error: You are already registered!")





def login():
    load_wb = load_workbook("#20$6&7!#04.xlsx", data_only=True)
    load_ws = load_wb['asdfas9df72']
    luserId_hash = hashlib.sha256(luser_id.get().encode())
    lpassword_hash = hashlib.sha256(lpassword.get().encode())
    if(luserId_hash.hexdigest() == load_ws['A1'].value and lpassword_hash.hexdigest() == load_ws['A2'].value):
        newWindow = Tk()
        newWindow.title("File Locker (Test Version: alphaV, Logined: " + luser_id.get() + ")")
        newWindow.geometry("600x300")
        newWindow.eval('tk::PlaceWindow . center')
        newWindow.configure(bg="#fff")
        window.destroy()
        standardLabel.stdLabel(newWindow, 85, 10, 18, 1, "#000", "#fff", "맑은 고딕", 30, "Choose Folder", "bold")
        standardLabel.stdLabel(newWindow, 0, 280, 15, 1, "#000", "#fff", "맑은 고딕", "8", "Logined: " + luser_id.get())
        standardLabel.stdLabel(newWindow, 500, 280, 15, 1, "#000", "#fff", "맑은 고딕", 8, "©2022 Gloomn")
        standardEntry.stdEntry(newWindow, 20, 120, 70, 1, "#a2a4a6", "#000", "#fff", "맑은 고딕", "10", "User ID", luser_id,'')
        newWindow.iconbitmap('resources/mainIcon_Black.ico')
        newWindow.mainloop()

    elif(len(luser_id.get()) == 0 or len(lpassword.get()) == 0):
        tkmb.showerror("Error Code: 004","Error: Please INSERT your ID or Password!")
    else:
        print("failed")
        print(luserId_hash.hexdigest())
        print(load_ws['A1'].value)
        tkmb.showerror("Error Code: 003", "Error: Please insert correct id & password!")
def test1():
    load_wb = load_workbook("#20$6&7!#04.xlsx", data_only=True)
    load_ws = load_wb['asdfas9df72']
    load_ws['A1'].value = 1
    load_ws['A2'].value = 1
    print(load_ws['A1'].value)
    print(load_ws['A2'].value)
    load_wb.save('#20$6&7!#04.xlsx')


def test2():
    a.insert(0, 'admin')

standardLabel.stdLabel(window, 85, 20, 18, 1, "#000","#fff", "맑은 고딕", 30, "File Locker Program", "bold")
standardLabel.stdLabel(window, 100, 115, 10, 1, "#000","#fff", "맑은 고딕", 13, "User Name:", "bold")
standardLabel.stdLabel(window, 110, 145, 10, 1, "#000","#fff", "맑은 고딕", 13, "Password:", "bold")
standardLabel.stdLabel(window, 500, 280, 15, 1, "#000","#fff", "맑은 고딕", 8, "©2022 Gloomn")
standardEntry.stdEntry(window,220, 120, 20,1,"#a2a4a6", "#000", "#fff", "맑은 고딕", "10", "User ID", luser_id, '')
standardEntry.stdEntry(window,220, 150, 20,1,"#a2a4a6", "#000", "#fff", "맑은 고딕", "10", "Password", lpassword, '*')
standardButton.fltButton(window, 220,189,20, 2,login, "Login", "#ffcc66", "#141414")
standardButton.fltButton(window, 220, 230, 20, 2, register, "Register", "#ffcc66", "#141414")
standardButton.fltButton(window, 0, 0, 5, 1, test1, "test1", "#ffcc66", "#141414")
window.bind('<Return>', lambda event:login())

#window main icon change
window.iconbitmap('resources/mainIcon_Black.ico')
window.mainloop()
